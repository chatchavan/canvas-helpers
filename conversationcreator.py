"""Send conversation messages to individual students on a course. Messages can be personalised or generic, and can also
include a unique attachment file."""

__author__ = 'Simon Robinson'
__copyright__ = 'Copyright (c) 2023 Simon Robinson'
__license__ = 'Apache 2.0'
__version__ = '2023-03-08'  # ISO 8601 (YYYY-MM-DD)

import argparse
import csv
import json
import mimetypes
import os

import openpyxl
import requests

from canvashelpers import Utils

parser = argparse.ArgumentParser()
parser.add_argument('url', nargs=1,
                    help='Please pass the URL of the course whose students will be sent comments')
parser.add_argument('--working-directory', default=None,
                    help='The root directory to use for the script\'s operation. Within this directory, attachments '
                         'and any `--comments-file` should be placed in a subfolder named as the course number (e.g., '
                         'for a course at https://[canvas-domain]/courses/10000/, name the subfolder 10000). Default: '
                         'the same directory as this script')
parser.add_argument('--attachment-extension', default='pdf',
                    help='The file extension of attachments to add to messages (without the dot separator). Files '
                         'should be named following the format [student number].[extension]. Default: \'pdf\'')
parser.add_argument('--attachment-mime-type', default=None,
                    help='Canvas requires a hint about the MIME type of the attachment file you are uploading. The '
                         'script is able to guess the correct value in most cases, but if you are uploading a file '
                         'with an unusual extension or format then you can specify a value here.')
parser.add_argument('--comments-file', default=None,
                    help='An XLSX or CSV file containing a minimum of one column: student number. A second column can '
                         'be added for per-student content that will be used as the conversation\'s message, '
                         'overriding the global `--conversation-message`.')
parser.add_argument('--conversation-subject', default='Course message',
                    help='The subject of the conversation. The default value is \'Course message\'')
parser.add_argument('--conversation-message', default='See attached file',
                    help='The conversation message to be sent. The default value is \'See attached file\', but this '
                         'can be overridden via this parameter or `--comments-file`. Use \\n for linebreaks')
parser.add_argument('--delete-after-sending', action='store_true',
                    help='Sending messages using this script can fill up your Sent folder. If that is an issue, use '
                         'this parameter to remove sent messages after sending. This only affects your own view of the '
                         'conversation; the message will remain in the recipient\'s Inbox, and if they reply you will '
                         'still see the original context')
parser.add_argument('--delete-conversation-attachments', action='store_true',
                    help='Sending many messages with attachments will quickly fill up the very small (52.4MB) default '
                         'Canvas per-user storage allowance. It is time-consuming to use the web interface to remove '
                         'files and restore space; instead, running the script with this parameter will remove *all* '
                         'files in your account\'s `conversation attachments` folder. If this parameter is set, all '
                         'others except `--dry-run` are ignored, and the script will exit after completion.')
parser.add_argument('--dry-run', action='store_true',
                    help='Preview the script\'s actions without actually making any changes. Highly recommended!')
args = parser.parse_args()  # exits if no assignment URL is provided

COURSE_URL = Utils.course_url_to_api(args.url[0])
COURSE_ID = Utils.get_course_id(COURSE_URL)
API_ROOT = COURSE_URL.split('/courses')[0]

# deleting files is a separate mode
if args.delete_conversation_attachments:
    folder_name = 'conversation attachments'
    print('DRY RUN:' if args.dry_run else '', 'removing all files from your `%s` folder' % folder_name)

    attachments_response = requests.get('%s/users/self/folders/by_path/%s' % (API_ROOT, folder_name),
                                        headers=Utils.canvas_api_headers())
    if attachments_response.status_code != 200:
        print('ERROR: unable to find your `%s` folder; aborting' % folder_name)
        exit()
    attachments_folder = attachments_response.json()[-1]  # resolve provides the requested folder last
    if attachments_folder['name'] != folder_name:
        print('ERROR: unable to match your `%s` folder; aborting' % folder_name)
        exit()

    folder_id = attachments_folder['id']
    user_files = Utils.canvas_multi_page_request('%s/users/self/files' % API_ROOT, type_hint='files')
    if not user_files:
        print('No files found in your user account; nothing to do')
        exit()

    user_files_json = json.loads(user_files)
    files_to_delete = []
    for file in user_files_json:
        if file['folder_id'] == folder_id:
            files_to_delete.append(file['id'])

    if len(files_to_delete) > 0:
        print('DRY RUN: would delete' if args.dry_run else 'Deleting', len(files_to_delete),
              'files from your `%s` folder' % folder_name)
        if args.dry_run:
            exit()
        for file_id in files_to_delete:
            delete_request = requests.delete('%s/files/%d' % (API_ROOT, file_id),
                                             headers=Utils.canvas_api_headers())
            if delete_request.status_code == 200:
                print('Deleted file', delete_request.text)
    else:
        print('No files found in your `%s` folder; nothing to do' % folder_name)
    exit()

INPUT_DIRECTORY = os.path.join(
    os.path.dirname(os.path.realpath(__file__)) if args.working_directory is None else args.working_directory,
    str(COURSE_ID))
if not os.path.exists(INPUT_DIRECTORY):
    print('ERROR: input directory not found - please place all files to upload (and any `--comments-file`) in the '
          'folder %s' % INPUT_DIRECTORY)
    exit()
print('%screating conversations for course %s' % ('DRY RUN: ' if args.dry_run else '', args.url[0]))

# load and parse comments
comments_map = {}
if args.comments_file is not None:
    comments_file = os.path.join(INPUT_DIRECTORY, args.comments_file)
    if os.path.exists(comments_file):
        if comments_file.lower().endswith('.xlsx'):
            comments_workbook = openpyxl.load_workbook(comments_file)
            comments_sheet = comments_workbook[comments_workbook.sheetnames[0]]
            for row in comments_sheet.iter_rows():
                comments_map[row[0].value] = row[1].value
        else:
            with open(comments_file, newline='') as marks_csv:
                reader = csv.reader(marks_csv)
                for row in reader:
                    comments_map[row[0]] = row[1]
        print('Loaded comments mapping for', len(comments_map), 'people:', comments_map)
    else:
        print('Ignoring comments file argument', args.comments_file, '- not found in course directory at',
              comments_file)

# get the course's students
course_user_response = Utils.get_course_users(COURSE_URL, enrolment_types=['student'])
if not course_user_response:
    print('ERROR: unable to retrieve course student list; aborting')
    exit()
course_user_json = json.loads(course_user_response)

SELF_ID, user_name = Utils.get_user_details(API_ROOT, user_id='self')
if not SELF_ID:
    print('ERROR: unable to retrieve your Canvas ID; aborting')
    exit()
# conversation attachments cannot be in sub-folders(!), but Canvas automatically handles duplicates (by renaming)
# FILES_SUBFOLDER_PATH = 'conversation attachments/%s/%d/%d' % (
#     os.path.splitext(os.path.basename(__file__))[0], COURSE_ID, int(time.time()))
FILES_SUBFOLDER_PATH = 'conversation attachments'
print('Generating', len(course_user_json), 'conversations and uploading attachments to', user_name, '\'s folder:',
      '%s/files/folder/users_%d/%s' % (args.url[0].split('/courses')[0], SELF_ID,
                                       FILES_SUBFOLDER_PATH.replace(' ', '%20')))  # display formatting only

for user in course_user_json:
    print('\nProcessing message to', user)

    canvas_id = user['id']
    student_number = user['login_id']

    attachment_file = '%s.%s' % (student_number, args.attachment_extension)
    attachment_path = os.path.join(INPUT_DIRECTORY, attachment_file)
    attachment_mime_type = args.attachment_mime_type if args.attachment_mime_type is not None else \
        mimetypes.guess_type(attachment_path)[0]

    attachment_exists = os.path.exists(attachment_path)
    if attachment_exists and attachment_mime_type is not None:
        print('Found conversation attachment file', attachment_file, 'with MIME type', attachment_mime_type)
    else:
        print('Attachment %s at %s %s; skipping upload for this conversation' % (
            attachment_file, attachment_path, ('not found' if not attachment_exists else
                                               'is not a recognised MIME type - see `--attachment-mime-type` option')))
        attachment_file = None

    # filter out unset fields, allowing any combination of mark/comment/attachment)
    conversation_message = args.conversation_message
    if student_number in comments_map:
        conversation_message = comments_map[student_number]
    elif attachment_file is None:
        print('Could not find attachment or message for conversation (at least one item is required); skipping')
        continue

    # see: https://canvas.instructure.com/doc/api/submissions.html#method.submissions_api.update
    conversation_data = {
        'recipients[]': [canvas_id],
        'subject': args.conversation_subject,
        'body': conversation_message.replace('\\n', '\n'),
        'force_new': True,
        'context_code': 'course_%d' % COURSE_ID
    }
    if conversation_message != args.conversation_message:
        print('Adding conversation message from spreadsheet:', conversation_message)
    else:
        print('Using conversation message provided as script argument:', conversation_message)

    if args.dry_run:
        print('DRY RUN: skipping attachment upload and message posting/deletion steps; moving to next recipient')
        continue

    if attachment_file is not None:
        # if there is an attachment we first need to request an upload URL, then associate with a submission comment
        submission_form_data = {
            'name': attachment_file,
            'content_type': attachment_mime_type,
            'parent_folder_path': FILES_SUBFOLDER_PATH
            # enable if multiple files need to be sent to the same people in a short period (by default, files with
            # clashing names are overwritten, and the old version shows as deleted to its original recipients)
            # 'on_duplicate': 'rename'
        }
        file_submission_url_response = requests.post('%s/users/self/files' % API_ROOT, data=submission_form_data,
                                                     headers=Utils.canvas_api_headers())
        if file_submission_url_response.status_code != 200:
            print('\tERROR: unable to retrieve attachment upload URL; skipping submission')
            continue

        file_submission_url_json = file_submission_url_response.json()
        print('\tUploading attachment to', file_submission_url_json['upload_url'].split('?')[0], '[truncated]')

        files_data = {'file': (attachment_file, open(attachment_path, 'rb'))}
        file_submission_upload_response = requests.post(file_submission_url_json['upload_url'],
                                                        data=submission_form_data, files=files_data,
                                                        headers=Utils.canvas_api_headers())

        if file_submission_upload_response.status_code != 201:  # note: 201 Created
            print('\tERROR: unable to upload attachment file; skipping recipient')
            continue

        file_submission_upload_json = file_submission_upload_response.json()
        print('\tAssociating uploaded file', file_submission_upload_json['id'], 'with conversation')
        conversation_data['attachment_ids[]'] = [file_submission_upload_json['id']]

    message_creation_response = requests.post('%s/conversations' % API_ROOT, data=conversation_data,
                                              headers=Utils.canvas_api_headers())
    if message_creation_response.status_code != 201:
        print('\tERROR: unable to send conversation message and/or associate attachment; skipping recipient')
        continue

    print('\tMessage successfully sent to user', canvas_id, '(%s)' % student_number)

    if args.delete_after_sending:
        sent_message = message_creation_response.json()
        message_deletion_response = requests.delete('%s/conversations/%d' % (API_ROOT, sent_message[0]['id']),
                                                    headers=Utils.canvas_api_headers())
        if message_deletion_response.status_code == 200:
            print('\tRemoved message from your Sent items folder')
        else:
            print('\tUnable to remove message from your sent items folder:', message_deletion_response.text)
